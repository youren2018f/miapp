import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
from PIL import Image

# 試著讀取圖片，若檔案不存在則跳過，避免程式崩潰
try:
    image = Image.open("k2.jpg")
except Exception:
    image = None

def get_week_day(day):
    week_day_dict = {
        0: "一",
        1: "二",
        2: "三",
        3: "四",
        4: "五",
        5: "六",
        6: "日"
    }
    return week_day_dict.get(day, "")

now = datetime.now()
chinese_weekday = get_week_day(now.weekday())

st.title('''
This app is mi-app
''')

st.write('''
***
''')
st.write("目前的受訪內容關鍵字：治安交通宣導、預防犯罪宣導、交通宣導、交通安全宣導、治安宣導")
st.header('上傳區')

agree = st.checkbox('是否已經看完Coding is Magic?')
st.write("[Coding is Magic](https://speakerdeck.com/mosky/coding-is-magic)")
if agree:
     st.write('棒棒喔!')

uploaded_file = st.file_uploader("請上傳節目受訪者申請表xlsx檔", type = "xlsx")

if uploaded_file is not None and agree:
    # 讀取上傳的檔案
    df = pd.read_excel(uploaded_file, usecols="A:M", header=1)

    # 過濾內容
    keywords = ["治安交通宣導", "預防犯罪宣導", "交通宣導", "交通安全宣導", "治安宣導"]
    df = df[df['受訪內容'].isin(keywords)]
 
    # 讀取範本檔 (請確保 111.xlsx 也在 GitHub 倉庫中)
    try:
        wb = openpyxl.load_workbook("111.xlsx")
        ws = wb['臺東分臺']
        
        # 填寫表頭資訊
        ws.cell(row=1, column=1).value = f'''警察廣播電臺警政宣導節目宣導記錄表
                                             {now.year - 1911}年{now.month}月{now.day}日（星期{chinese_weekday}）臺東分臺 潘亭羽 製表'''
        
        # 填寫資料內容
        i = 4
        for ind in df.index:
            ws.cell(row=i, column=1).value = df.loc[ind, "日期"]
            ws.cell(row=i, column=2).value = df.loc[ind, "時間"]
            ws.cell(row=i, column=3).value = "臺東分臺\n（FM94.3）"
            ws.cell(row=i, column=4).value = df.loc[ind, "節目名稱"]
            ws.cell(row=i, column=5).value = df.loc[ind, "主持人"]
            ws.cell(row=i, column=6).value = df.loc[ind, "單位"]
            ws.cell(row=i, column=7).value = df.loc[ind, "職稱"]
            ws.cell(row=i, column=8).value = df.loc[ind, "受訪者"]
            ws.cell(row=i, column=9).value = df.loc[ind, "受訪內容"]
            ws.cell(row=i, column=10).value = "Facebook粉絲專頁發文、官方網站專區宣導、各警分局粉絲專頁連結警廣警政專訪網址分享"
            i = i + 1

        # --- 核心修正：取代 save_virtual_workbook ---
        output = BytesIO()
        wb.save(output)
        data = output.getvalue()
        # ------------------------------------------

        st.write('''
        ***
        ''')
        st.header('下載區')
        st.download_button(
            label="下載檔案",
            data=data,
            file_name="yoyo_of_file.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        if image:
            st.image(image, caption='阿奇小狗狗')
            
    except FileNotFoundError:
        st.error("找不到範本檔案 '111.xlsx'，請確認該檔案已上傳至 GitHub。")
    except Exception as e:
        st.error(f"發生錯誤：{e}")
